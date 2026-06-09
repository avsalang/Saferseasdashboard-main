from __future__ import annotations

import json
import math
import subprocess
import sys
import tempfile
import time
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.properties import CalcProperties
from openpyxl.worksheet.table import Table, TableStyleInfo


REPO_ROOT = Path(__file__).resolve().parents[1]
OUTPUT_DIR = REPO_ROOT / "outputs"
OUTPUT_PATH = OUTPUT_DIR / "SAFERSEAS_Analytics_Implementation_Guide.xlsx"

TITLE_FILL = PatternFill("solid", fgColor="1E3A8A")
SECTION_FILL = PatternFill("solid", fgColor="DBEAFE")
SUBSECTION_FILL = PatternFill("solid", fgColor="EFF6FF")
CARD_FILL = PatternFill("solid", fgColor="FFFFFF")
NOTE_FILL = PatternFill("solid", fgColor="F8FAFC")
WARN_FILL = PatternFill("solid", fgColor="FEF2F2")
THIN_BORDER = Border(
    left=Side(style="thin", color="D1D5DB"),
    right=Side(style="thin", color="D1D5DB"),
    top=Side(style="thin", color="D1D5DB"),
    bottom=Side(style="thin", color="D1D5DB"),
)

RAW_HEADERS = [
    "Incident ID",
    "Incident DateTime",
    "Incident Type",
    "Casualty Type",
    "Severity",
    "Vessel Name",
    "Former Names",
    "Vessel Type",
    "Flag State",
    "IMO Number",
    "Official Number",
    "Gross Tonnage",
    "Length Meters",
    "Breadth Meters",
    "Propulsion Type",
    "Engine Power kW",
    "Classification Society",
    "Year Built",
    "Latitude",
    "Longitude",
    "Reference Port",
    "Province",
    "Sea Area",
    "Corridor",
    "Wave Height (m)",
    "Wind Force",
    "Visibility",
    "Owner Name",
    "Owner Contact",
    "Operator Name",
    "Operator Address",
    "Voyage Origin",
    "Voyage Destination",
    "Intended Route",
    "Cargo Type",
    "Cargo Quantity",
    "Authority",
    "Reporting Method",
    "Status",
    "Narrative",
    "Key Events Timeline",
    "Crew Response",
    "Crew Behavior Factors",
    "Primary Cause",
    "Findings Of Cause",
    "Damages Incurred",
    "Crew Count",
    "Passenger Count",
    "Fatalities",
    "Injuries",
    "Recommendations",
    "Linked Documents",
]

LOOKUP_PROVINCES = [
    ("Metro Manila", "Luzon"),
    ("Batangas", "Luzon"),
    ("Cebu", "Visayas"),
    ("Bohol", "Visayas"),
    ("Albay", "Luzon"),
    ("Davao del Sur", "Mindanao"),
    ("Leyte", "Visayas"),
    ("Southern Leyte", "Visayas"),
    ("Palawan", "Luzon"),
    ("Cagayan", "Luzon"),
    ("Oriental Mindoro", "Luzon"),
    ("Misamis Oriental", "Mindanao"),
    ("Sorsogon", "Luzon"),
    ("Iloilo", "Visayas"),
]

LOOKUP_FALLBACK_CAUSE = [
    ("Collision", "Navigation Error"),
    ("Grounding", "Navigation Error"),
    ("Sinking", "Weather Conditions"),
    ("Fire/Explosion", "Mechanical Failure"),
    ("Capsizing", "Weather Conditions"),
    ("Flooding", "Mechanical Failure"),
    ("Swamping", "Weather Conditions"),
    ("Machinery Failure", "Mechanical Failure"),
    ("Person Overboard", "Human Error"),
    ("Pollution Incident", "Mechanical Failure"),
]

LOOKUP_BEHAVIOR = [
    ("Navigation Error", "Communication Failure | Violation of Procedures"),
    ("Mechanical Failure", "Inadequate Training"),
    ("Weather Conditions", "Communication Failure | Violation of Procedures"),
    ("Human Error", "Fatigue | Violation of Procedures"),
    ("Under Investigation", ""),
]

LOOKUP_CARGO = [
    ("Ferry", "Passengers"),
    ("Fishing Vessel", "Fish Catch"),
    ("Container Ship", "Containers"),
    ("Oil Tanker", "Fuel / Oil"),
    ("Bulk Carrier", "Bulk Cargo"),
    ("General Cargo", "General Cargo"),
    ("Passenger Vessel", "Passengers"),
]

LOOKUP_YEAR_BUILT = [
    ("Ferry", 2009),
    ("Fishing Vessel", 2001),
    ("Container Ship", 2010),
    ("Oil Tanker", 2007),
    ("Bulk Carrier", 2005),
    ("General Cargo", 2008),
    ("Passenger Vessel", 2012),
]

LOOKUP_SEVERITY = [
    ("Very Serious", 4, "#991B1B"),
    ("Serious", 3, "#DC2626"),
    ("Less Serious", 2, "#F97316"),
    ("Near Miss", 1, "#F59E0B"),
]

SEASON_ORDER = ["Northeast Monsoon", "Transition Season", "Southwest Monsoon", "Typhoon Season"]
VISIBILITY_ORDER = ["Good", "Moderate", "Poor", "Very Poor"]


def run_export(temp_dir: Path) -> tuple[list[dict], dict]:
    exporter_js = temp_dir / "export_runtime_data.mjs"
    exporter_js.write_text(
        """
import fs from "node:fs";
import path from "node:path";
import { createRequire } from "node:module";

const repoRoot = process.argv[2];
const outputDir = process.argv[3];
const require = createRequire(path.join(repoRoot, "package.json"));
const { buildSync } = require("esbuild");

const mockOut = path.join(outputDir, "mockIncidents.cjs");
const insightsOut = path.join(outputDir, "incidentInsights.cjs");

buildSync({
  entryPoints: [path.join(repoRoot, "src/app/data/mockIncidents.ts")],
  bundle: true,
  platform: "node",
  format: "cjs",
  outfile: mockOut,
});

buildSync({
  entryPoints: [path.join(repoRoot, "src/app/data/incidentInsights.ts")],
  bundle: true,
  platform: "node",
  format: "cjs",
  outfile: insightsOut,
});

const mockModule = require(mockOut);
const insightsModule = require(insightsOut);

const insightKeys = [
  "analyticsSummary",
  "monthlyTrendData",
  "severityBreakdown",
  "corridorHotspots",
  "seaAreaBreakdown",
  "causeParetoData",
  "casualtiesByType",
  "behaviorFactorBreakdown",
  "seasonBreakdown",
  "seasonVisibilityMatrix",
  "seasonVisibilityPeak",
  "vesselExposureData",
  "dataCompleteness",
  "criticalGapData",
  "statusCompletenessData",
  "reportingMethodBreakdown",
  "policyEvidencePackets",
  "corridorInvestigationTable",
];

const insights = {};
for (const key of insightKeys) {
  insights[key] = insightsModule[key];
}

fs.writeFileSync(path.join(outputDir, "mockIncidents.json"), JSON.stringify(mockModule.mockIncidents, null, 2));
fs.writeFileSync(path.join(outputDir, "incidentInsights.json"), JSON.stringify(insights, null, 2));
""".strip(),
        encoding="utf-8",
    )

    subprocess.run(
        ["node", str(exporter_js), str(REPO_ROOT), str(temp_dir)],
        check=True,
        cwd=REPO_ROOT,
    )

    mock_incidents = json.loads((temp_dir / "mockIncidents.json").read_text(encoding="utf-8"))
    insights = json.loads((temp_dir / "incidentInsights.json").read_text(encoding="utf-8"))
    return mock_incidents, insights


def iso_to_excel_datetime(value: str) -> datetime:
    return datetime.fromisoformat(value.replace("Z", "+00:00")).replace(tzinfo=None)


def flatten(value):
    if value is None:
        return ""
    if isinstance(value, list):
        return " | ".join(str(item) for item in value)
    return value


def apply_title(ws, title: str, subtitle: str | None = None, width: int = 9):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width)
    cell = ws.cell(1, 1, title)
    cell.fill = TITLE_FILL
    cell.font = Font(color="FFFFFF", bold=True, size=16)
    cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=width)
        sub = ws.cell(2, 1, subtitle)
        sub.fill = SUBSECTION_FILL
        sub.font = Font(color="334155", italic=True, size=10)
        sub.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[2].height = 20


def style_section_header(ws, row: int, title: str, width: int = 6):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=width)
    cell = ws.cell(row, 1, title)
    cell.fill = SECTION_FILL
    cell.font = Font(color="1E3A8A", bold=True, size=11)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20


def style_table_header(ws, row: int, start_col: int, headers: list[str]):
    for offset, header in enumerate(headers):
        cell = ws.cell(row, start_col + offset, header)
        cell.fill = SECTION_FILL
        cell.font = Font(color="1E3A8A", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_range(ws, min_row: int, max_row: int, min_col: int, max_col: int, number_formats: dict[int, str] | None = None):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            if cell.row > min_row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            if number_formats and cell.column in number_formats and cell.row > min_row:
                cell.number_format = number_formats[cell.column]


def set_widths(ws, widths: dict[str, float]):
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def add_excel_table(ws, ref: str, name: str):
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def add_card(ws, start_cell: str, title: str, value_formula: str, subtitle: str, fill_color: str):
    col = ws[start_cell].column
    row = ws[start_cell].row
    ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 1)
    ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 2, end_column=col + 1)
    ws.merge_cells(start_row=row + 3, start_column=col, end_row=row + 4, end_column=col + 1)

    title_cell = ws.cell(row, col, title)
    title_cell.fill = PatternFill("solid", fgColor=fill_color)
    title_cell.font = Font(color="FFFFFF", bold=True, size=10)
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    value_cell = ws.cell(row + 1, col, value_formula)
    value_cell.fill = CARD_FILL
    value_cell.font = Font(color=fill_color, bold=True, size=18)
    value_cell.alignment = Alignment(horizontal="center", vertical="center")

    note_cell = ws.cell(row + 3, col, subtitle)
    note_cell.fill = NOTE_FILL
    note_cell.font = Font(color="475569", size=9)
    note_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in range(row, row + 5):
        for c in range(col, col + 2):
            ws.cell(r, c).border = THIN_BORDER

    for offset in range(5):
        ws.row_dimensions[row + offset].height = 18 if offset not in (1, 2) else 22


def create_chart_data_titles(chart, color: str):
    if chart.ser:
        series = chart.ser[0]
        series.graphicalProperties.solidFill = color.replace("#", "")
        series.graphicalProperties.line.solidFill = color.replace("#", "")


def build_workbook(mock_incidents: list[dict], insights: dict) -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)

    guide = wb.create_sheet("Guide")
    raw_ws = wb.create_sheet("Raw_Incidents")
    lookup_ws = wb.create_sheet("Lookups")
    enriched_ws = wb.create_sheet("Enriched")
    trend_ws = wb.create_sheet("Trend_Summary")
    causal_ws = wb.create_sheet("Causal_Analysis")
    exposure_ws = wb.create_sheet("Exposure_Analysis")
    quality_ws = wb.create_sheet("Quality_Analysis")
    policy_ws = wb.create_sheet("Policy_Brief")

    apply_title(
        guide,
        "SAFERSEAS Analytics Implementation Guide",
        "Formula-driven Excel companion to the mockup using the same dummy dataset and equivalent transferable visuals.",
        width=10,
    )
    guide["A4"] = "Purpose"
    guide["A4"].fill = SECTION_FILL
    guide["A4"].font = Font(color="1E3A8A", bold=True)
    guide["A5"] = (
        "This workbook mirrors the mockup's analytical outputs so a developer can inspect raw data, derived logic, chart helper tables, "
        "and policy evidence calculations in Excel."
    )
    guide["A5"].alignment = Alignment(wrap_text=True)
    guide["A7"] = "Included workbook sections"
    guide["A7"].fill = SECTION_FILL
    guide["A7"].font = Font(color="1E3A8A", bold=True)
    included = [
        "Raw_Incidents: source dummy data from the mockup",
        "Lookups: region maps, fallback logic, season order, severity weights",
        "Enriched: formula-driven derived fields used by analytics",
        "Trend / Causal / Exposure / Quality sheets: chart helper tables plus Excel charts",
        "Policy_Brief: policy packets, corridor register, data caveats, methodology notes",
    ]
    for idx, item in enumerate(included, start=8):
        guide[f"A{idx}"] = f"- {item}"
    guide["A15"] = "Not transferred from the mockup"
    guide["A15"].fill = WARN_FILL
    guide["A15"].font = Font(color="991B1B", bold=True)
    guide["A16"] = (
        "The interactive GIS density/cluster map is not reproduced as a native Excel visual. "
        "The workbook retains the same underlying coordinates and corridor logic, but not the map interaction layer."
    )
    guide["A16"].alignment = Alignment(wrap_text=True)
    guide["A18"] = "Method note"
    guide["A18"].fill = SECTION_FILL
    guide["A18"].font = Font(color="1E3A8A", bold=True)
    guide["A19"] = (
        "Risk Index and Confidence Score are included because the mockup uses them, but both remain working proxies for implementation guidance, "
        "not validated operational doctrine."
    )
    guide["A19"].alignment = Alignment(wrap_text=True)
    set_widths(guide, {"A": 38, "B": 22, "C": 18, "D": 18, "E": 18, "F": 18, "G": 18, "H": 18, "I": 18, "J": 18})

    apply_title(raw_ws, "Raw Incident Dataset", "Direct export of the current SAFERSEAS dummy dataset.", width=12)
    style_table_header(raw_ws, 4, 1, RAW_HEADERS)
    raw_ws.freeze_panes = "A5"
    raw_rows = []
    for incident in mock_incidents:
        raw_rows.append([
            incident["id"],
            iso_to_excel_datetime(incident["date"]),
            incident["type"],
            incident["casualtyType"],
            incident["severity"],
            incident["vesselName"],
            flatten(incident.get("formerNames")),
            incident["vesselType"],
            incident["flagState"],
            incident.get("imoNumber", ""),
            incident.get("officialNumber", ""),
            incident.get("grossTonnage", ""),
            incident.get("lengthMeters", ""),
            incident.get("breadthMeters", ""),
            incident.get("propulsionType", ""),
            incident.get("enginePowerKw", ""),
            incident.get("classificationSociety", ""),
            incident.get("yearBuilt", ""),
            incident["location"]["lat"],
            incident["location"]["lng"],
            incident["location"]["port"],
            incident["location"]["province"],
            incident.get("seaArea", ""),
            incident.get("corridor", ""),
            incident["weather"]["waveHeight"],
            incident["weather"]["windForce"],
            incident["weather"]["visibility"],
            incident.get("ownerName", ""),
            incident.get("ownerContact", ""),
            incident.get("operatorName", ""),
            incident.get("operatorAddress", ""),
            incident.get("voyageOrigin", ""),
            incident.get("voyageDestination", ""),
            incident.get("intendedRoute", ""),
            incident.get("cargoType", ""),
            incident.get("cargoQuantity", ""),
            incident["authority"],
            incident.get("reportingMethod", ""),
            incident["status"],
            incident["narrative"],
            incident.get("keyEventsTimeline", ""),
            incident.get("crewResponse", ""),
            flatten(incident.get("crewBehaviorFactors")),
            incident.get("primaryCause", ""),
            incident.get("findingsOfCause", ""),
            flatten(incident.get("damagesIncurred")),
            incident["crewCount"],
            incident["passengerCount"],
            incident["fatalities"],
            incident["injuries"],
            incident.get("recommendations", ""),
            flatten(incident.get("linkedDocuments")),
        ])
    for r_idx, row in enumerate(raw_rows, start=5):
        for c_idx, value in enumerate(row, start=1):
            raw_ws.cell(r_idx, c_idx, value)
    style_data_range(
        raw_ws,
        4,
        4 + len(raw_rows),
        1,
        len(RAW_HEADERS),
        number_formats={2: "yyyy-mm-dd hh:mm", 19: "0.0000", 20: "0.0000", 25: "0.0", 26: "0"},
    )
    add_excel_table(raw_ws, f"A4:AZ{4 + len(raw_rows)}", "RawIncidents")
    set_widths(
        raw_ws,
        {
            "A": 18, "B": 20, "C": 16, "D": 16, "E": 14, "F": 22, "G": 20, "H": 18, "I": 14, "J": 16,
            "K": 16, "L": 14, "M": 14, "N": 14, "O": 18, "P": 14, "Q": 18, "R": 12, "S": 12, "T": 12,
            "U": 18, "V": 16, "W": 18, "X": 28, "Y": 14, "Z": 12, "AA": 14, "AB": 22, "AC": 18, "AD": 22,
            "AE": 24, "AF": 16, "AG": 18, "AH": 28, "AI": 18, "AJ": 22, "AK": 14, "AL": 16, "AM": 14,
            "AN": 34, "AO": 32, "AP": 32, "AQ": 28, "AR": 18, "AS": 34, "AT": 22, "AU": 12, "AV": 14,
            "AW": 12, "AX": 12, "AY": 34, "AZ": 28,
        },
    )

    apply_title(lookup_ws, "Lookup Tables", "Model assumptions and categorical mappings mirrored from the mockup logic.", width=8)
    sections = [
        ("A4", "Province to Region", ["Province", "Region"], LOOKUP_PROVINCES),
        ("D4", "Fallback Cause by Incident Type", ["Incident Type", "Fallback Cause"], LOOKUP_FALLBACK_CAUSE),
        ("G4", "Fallback Behavior by Cause", ["Cause", "Fallback Behavior"], LOOKUP_BEHAVIOR),
        ("J4", "Fallback Cargo by Vessel Type", ["Vessel Type", "Fallback Cargo"], LOOKUP_CARGO),
        ("M4", "Fallback Year Built by Vessel Type", ["Vessel Type", "Fallback Year Built"], LOOKUP_YEAR_BUILT),
        ("P4", "Severity Weights", ["Severity", "Weight", "Color"], LOOKUP_SEVERITY),
    ]
    for anchor, title, headers, rows in sections:
        start_col = lookup_ws[anchor].column
        start_row = lookup_ws[anchor].row
        lookup_ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + len(headers) - 1)
        cell = lookup_ws.cell(start_row, start_col, title)
        cell.fill = SECTION_FILL
        cell.font = Font(color="1E3A8A", bold=True)
        style_table_header(lookup_ws, start_row + 1, start_col, headers)
        for r_offset, values in enumerate(rows, start=2):
            for c_offset, value in enumerate(values):
                lookup_ws.cell(start_row + r_offset, start_col + c_offset, value)
        style_data_range(lookup_ws, start_row + 1, start_row + 1 + len(rows), start_col, start_col + len(headers) - 1)
    lookup_ws["T4"] = "Season Order"
    lookup_ws["T4"].fill = SECTION_FILL
    lookup_ws["T4"].font = Font(color="1E3A8A", bold=True)
    for idx, value in enumerate(SEASON_ORDER, start=5):
        lookup_ws[f"T{idx}"] = value
    lookup_ws["U4"] = "Visibility Order"
    lookup_ws["U4"].fill = SECTION_FILL
    lookup_ws["U4"].font = Font(color="1E3A8A", bold=True)
    for idx, value in enumerate(VISIBILITY_ORDER, start=5):
        lookup_ws[f"U{idx}"] = value
    set_widths(lookup_ws, {"A": 18, "B": 12, "D": 18, "E": 18, "G": 18, "H": 28, "J": 18, "K": 18, "M": 18, "N": 18, "P": 16, "Q": 10, "R": 12, "T": 20, "U": 14})

    apply_title(
        enriched_ws,
        "Enriched Incident Model",
        "Formula-driven transformation layer matching the mockup's derived analytics fields.",
        width=10,
    )
    enriched_headers = [
        "Incident ID", "Incident DateTime", "Incident Type", "Severity", "Vessel Type", "Authority", "Status",
        "Province", "Sea Area Resolved", "Corridor Resolved", "Region", "Visibility", "Month Label", "Month Key",
        "Year", "Season", "Primary Cause Resolved", "Crew Behavior Factors Resolved", "Cargo Type Resolved",
        "Vessel Age Years", "Fatalities", "Injuries", "Casualty Count", "Severity Weight", "Risk Index",
        "Registry ID Captured", "Voyage Captured", "Weather Captured", "Route Label", "Standardization Score",
        "Linked Evidence Flag", "Review Matured Flag", "Narrative", "Vessel Name", "Latitude", "Longitude", "Reporting Method",
    ]
    style_table_header(enriched_ws, 4, 1, enriched_headers)
    enriched_ws.freeze_panes = "A5"
    raw_last = 4 + len(raw_rows)
    for idx in range(5, raw_last + 1):
        rr = idx
        enriched_ws.cell(idx, 1, f"=Raw_Incidents!A{rr}")
        enriched_ws.cell(idx, 2, f"=Raw_Incidents!B{rr}")
        enriched_ws.cell(idx, 3, f"=Raw_Incidents!C{rr}")
        enriched_ws.cell(idx, 4, f"=Raw_Incidents!E{rr}")
        enriched_ws.cell(idx, 5, f"=Raw_Incidents!H{rr}")
        enriched_ws.cell(idx, 6, f"=Raw_Incidents!AK{rr}")
        enriched_ws.cell(idx, 7, f"=Raw_Incidents!AM{rr}")
        enriched_ws.cell(idx, 8, f"=Raw_Incidents!V{rr}")
        enriched_ws.cell(idx, 9, f'=IF(Raw_Incidents!W{rr}<>"",Raw_Incidents!W{rr},Raw_Incidents!U{rr})')
        enriched_ws.cell(idx, 10, f'=IF(Raw_Incidents!X{rr}<>"",Raw_Incidents!X{rr},IF(Raw_Incidents!AH{rr}<>"",Raw_Incidents!AH{rr},IF(AND(Raw_Incidents!AF{rr}<>"",Raw_Incidents!AG{rr}<>""),Raw_Incidents!AF{rr}&" to "&Raw_Incidents!AG{rr},"Corridor pending")))')
        enriched_ws.cell(idx, 11, f'=IFERROR(INDEX(Lookups!$B$6:$B$19,MATCH(Raw_Incidents!V{rr},Lookups!$A$6:$A$19,0)),"National")')
        enriched_ws.cell(idx, 12, f"=Raw_Incidents!AA{rr}")
        enriched_ws.cell(idx, 13, f'=TEXT(B{rr},"mmm yy")')
        enriched_ws.cell(idx, 14, f'=TEXT(B{rr},"yyyy-mm")')
        enriched_ws.cell(idx, 15, f"=YEAR(B{rr})")
        enriched_ws.cell(idx, 16, f'=IF(OR(MONTH(B{rr})=12,MONTH(B{rr})<=2),"Northeast Monsoon",IF(AND(MONTH(B{rr})>=3,MONTH(B{rr})<=5),"Transition Season",IF(AND(MONTH(B{rr})>=6,MONTH(B{rr})<=9),"Southwest Monsoon","Typhoon Season")))')
        enriched_ws.cell(idx, 17, f'=IF(Raw_Incidents!AR{rr}<>"",Raw_Incidents!AR{rr},IFERROR(INDEX(Lookups!$E$6:$E$15,MATCH(C{rr},Lookups!$D$6:$D$15,0)),"Under Investigation"))')
        enriched_ws.cell(idx, 18, f'=IF(Raw_Incidents!AQ{rr}<>"",Raw_Incidents!AQ{rr},IFERROR(INDEX(Lookups!$H$6:$H$10,MATCH(Q{rr},Lookups!$G$6:$G$10,0)),""))')
        enriched_ws.cell(idx, 19, f'=IF(Raw_Incidents!AI{rr}<>"",Raw_Incidents!AI{rr},IFERROR(INDEX(Lookups!$K$6:$K$12,MATCH(E{rr},Lookups!$J$6:$J$12,0)),"General Cargo"))')
        enriched_ws.cell(idx, 20, f'=MAX(YEAR(TODAY())-IF(Raw_Incidents!R{rr}<>"",Raw_Incidents!R{rr},IFERROR(INDEX(Lookups!$N$6:$N$12,MATCH(E{rr},Lookups!$M$6:$M$12,0)),2010)),0)')
        enriched_ws.cell(idx, 21, f"=Raw_Incidents!AW{rr}")
        enriched_ws.cell(idx, 22, f"=Raw_Incidents!AX{rr}")
        enriched_ws.cell(idx, 23, f"=U{rr}+V{rr}")
        enriched_ws.cell(idx, 24, f'=IFERROR(INDEX(Lookups!$Q$6:$Q$9,MATCH(D{rr},Lookups!$P$6:$P$9,0)),1)')
        enriched_ws.cell(idx, 25, f"=X{rr}*MAX(W{rr},1)")
        enriched_ws.cell(idx, 26, f'=--OR(Raw_Incidents!J{rr}<>"",Raw_Incidents!K{rr}<>"")')
        enriched_ws.cell(idx, 27, f'=--AND(Raw_Incidents!AF{rr}<>"",Raw_Incidents!AG{rr}<>"")')
        enriched_ws.cell(idx, 28, f'=--AND(ISNUMBER(Raw_Incidents!Y{rr}),ISNUMBER(Raw_Incidents!Z{rr}),Raw_Incidents!AA{rr}<>"")')
        enriched_ws.cell(idx, 29, f'=IF(Raw_Incidents!AH{rr}<>"",Raw_Incidents!AH{rr},IF(AND(Raw_Incidents!AF{rr}<>"",Raw_Incidents!AG{rr}<>""),Raw_Incidents!AF{rr}&" to "&Raw_Incidents!AG{rr},"Route pending"))')
        enriched_ws.cell(idx, 30, f'=ROUND((--(B{rr}<>"")+--(C{rr}<>"")+--(D{rr}<>"")+--(Raw_Incidents!S{rr}<>"")+--(Raw_Incidents!T{rr}<>"")+--(Raw_Incidents!U{rr}<>"")+--(H{rr}<>"")+AB{rr}+--(Raw_Incidents!F{rr}<>"")+--(Raw_Incidents!I{rr}<>"")+Z{rr}+AA{rr}+--(Raw_Incidents!AI{rr}<>"")+--(Raw_Incidents!AN{rr}<>"")+--(F{rr}<>""))/15*100,0)')
        enriched_ws.cell(idx, 31, f'=--(Raw_Incidents!AZ{rr}<>"")')
        enriched_ws.cell(idx, 32, f'=--OR(G{rr}="Verified",G{rr}="Published")')
        enriched_ws.cell(idx, 33, f"=Raw_Incidents!AN{rr}")
        enriched_ws.cell(idx, 34, f"=Raw_Incidents!F{rr}")
        enriched_ws.cell(idx, 35, f"=Raw_Incidents!S{rr}")
        enriched_ws.cell(idx, 36, f"=Raw_Incidents!T{rr}")
        enriched_ws.cell(idx, 37, f"=Raw_Incidents!AL{rr}")
    style_data_range(
        enriched_ws,
        4,
        raw_last,
        1,
        len(enriched_headers),
        number_formats={2: "yyyy-mm-dd hh:mm", 20: "0", 21: "0", 22: "0", 23: "0", 24: "0", 25: "0", 30: "0%"},
    )
    add_excel_table(enriched_ws, f"A4:AK{raw_last}", "EnrichedIncidents")
    set_widths(enriched_ws, {"A": 18, "B": 20, "C": 16, "D": 14, "E": 18, "F": 14, "G": 14, "H": 16, "I": 18, "J": 28, "K": 14, "L": 12, "M": 12, "N": 12, "O": 10, "P": 20, "Q": 18, "R": 28, "S": 18, "T": 12, "U": 10, "V": 10, "W": 12, "X": 12, "Y": 10, "Z": 12, "AA": 12, "AB": 12, "AC": 28, "AD": 18, "AE": 12, "AF": 12, "AG": 34, "AH": 20, "AI": 12, "AJ": 12, "AK": 16})

    def setup_sheet(ws, title, subtitle):
        apply_title(ws, title, subtitle, width=16)
        ws.sheet_view.showGridLines = False

    setup_sheet(trend_ws, "Trend and Summary Analytics", "Excel equivalent of the mockup's overview and trend visuals.")
    add_card(trend_ws, "A4", "Incident Records", "=COUNTA(Enriched!$A$5:$A$999)", "Historical files in the evidence base", "1D4ED8")
    add_card(trend_ws, "C4", "High-Consequence Cases", '=SUMPRODUCT(--(((Enriched!$D$5:$D$999="Very Serious")+(Enriched!$W$5:$W$999>=3))>0))', "Very serious or multi-casualty cases", "991B1B")
    add_card(
        trend_ws,
        "E4",
        "Maritime Corridors",
        '=SUMPRODUCT((Enriched!$J$5:$J$999<>"")/COUNTIF(Enriched!$J$5:$J$999,Enriched!$J$5:$J$999&""))',
        "Distinct corridors represented in the model",
        "0F766E",
    )
    add_card(trend_ws, "G4", "Linked Evidence Coverage", '=ROUND(COUNTIF(Enriched!$AE$5:$AE$999,1)/COUNTA(Enriched!$A$5:$A$999)*100,0)&"%"', "Records with cited attachments or evidence", "7C3AED")

    monthly = insights["monthlyTrendData"]
    style_section_header(trend_ws, 11, "Monthly Incident and Casualty Trend", width=7)
    mt_headers = ["Period", "Incidents", "Fatalities", "Injuries", "Casualties", "Risk Index", "Very Serious"]
    style_table_header(trend_ws, 12, 1, mt_headers)
    for idx, row in enumerate(monthly, start=13):
        trend_ws.cell(idx, 1, row["period"])
        trend_ws.cell(idx, 2, f'=COUNTIF(Enriched!$M$5:$M${raw_last},"{row["period"]}")')
        trend_ws.cell(idx, 3, f'=SUMIF(Enriched!$M$5:$M${raw_last},"{row["period"]}",Enriched!$U$5:$U${raw_last})')
        trend_ws.cell(idx, 4, f'=SUMIF(Enriched!$M$5:$M${raw_last},"{row["period"]}",Enriched!$V$5:$V${raw_last})')
        trend_ws.cell(idx, 5, f"=C{idx}+D{idx}")
        trend_ws.cell(idx, 6, f'=SUMIF(Enriched!$M$5:$M${raw_last},"{row["period"]}",Enriched!$Y$5:$Y${raw_last})')
        trend_ws.cell(idx, 7, f'=COUNTIFS(Enriched!$M$5:$M${raw_last},"{row["period"]}",Enriched!$D$5:$D${raw_last},"Very Serious")')
    style_data_range(trend_ws, 12, 12 + len(monthly), 1, 7)

    chart = BarChart()
    chart.type = "col"
    chart.title = "Monthly Incident and Casualty Trend"
    chart.y_axis.title = "Incident Count"
    chart.height = 8.5
    chart.width = 11.5
    chart.add_data(Reference(trend_ws, min_col=2, max_col=2, min_row=12, max_row=12 + len(monthly)), titles_from_data=True)
    chart.set_categories(Reference(trend_ws, min_col=1, min_row=13, max_row=12 + len(monthly)))
    chart.legend.position = "r"
    create_chart_data_titles(chart, "#1D4ED8")
    line = LineChart()
    line.y_axis.title = "Casualties"
    line.y_axis.axId = 200
    line.y_axis.crosses = "max"
    line.add_data(Reference(trend_ws, min_col=3, max_col=4, min_row=12, max_row=12 + len(monthly)), titles_from_data=True)
    line.set_categories(Reference(trend_ws, min_col=1, min_row=13, max_row=12 + len(monthly)))
    for series, color in zip(line.ser, ["991B1B", "D97706"]):
        series.graphicalProperties.line.solidFill = color
        series.graphicalProperties.solidFill = color
        series.marker.symbol = "circle"
    chart += line
    trend_ws.add_chart(chart, "I12")

    style_section_header(trend_ws, 28, "Consequence Profile by Severity", width=6)
    sev_headers = ["Severity", "Records", "Percentage", "Casualties"]
    style_table_header(trend_ws, 29, 1, sev_headers)
    for idx, row in enumerate(insights["severityBreakdown"], start=30):
        trend_ws.cell(idx, 1, row["label"])
        trend_ws.cell(idx, 2, f'=COUNTIF(Enriched!$D$5:$D${raw_last},"{row["label"]}")')
        trend_ws.cell(idx, 3, f'=ROUND(B{idx}/COUNTA(Enriched!$A$5:$A${raw_last})*100,0)')
        trend_ws.cell(idx, 4, f'=SUMIF(Enriched!$D$5:$D${raw_last},"{row["label"]}",Enriched!$W$5:$W${raw_last})')
    style_data_range(trend_ws, 29, 29 + len(insights["severityBreakdown"]), 1, 4)
    sev_chart = BarChart()
    sev_chart.type = "col"
    sev_chart.title = "Consequence Profile by Severity"
    sev_chart.height = 8
    sev_chart.width = 11.5
    sev_chart.add_data(Reference(trend_ws, min_col=2, max_col=2, min_row=29, max_row=29 + len(insights["severityBreakdown"])), titles_from_data=True)
    sev_chart.set_categories(Reference(trend_ws, min_col=1, min_row=30, max_row=29 + len(insights["severityBreakdown"])))
    severity_colors = [row["color"].replace("#", "") for row in insights["severityBreakdown"]]
    for series in sev_chart.ser:
        series.graphicalProperties.solidFill = "1D4ED8"
        series.graphicalProperties.line.solidFill = "1D4ED8"
    sev_line = LineChart()
    sev_line.y_axis.axId = 201
    sev_line.y_axis.crosses = "max"
    sev_line.add_data(Reference(trend_ws, min_col=4, max_col=4, min_row=29, max_row=29 + len(insights["severityBreakdown"])), titles_from_data=True)
    sev_line.set_categories(Reference(trend_ws, min_col=1, min_row=30, max_row=29 + len(insights["severityBreakdown"])))
    sev_line.ser[0].graphicalProperties.line.solidFill = "475569"
    sev_line.ser[0].marker.symbol = "diamond"
    sev_chart += sev_line
    trend_ws.add_chart(sev_chart, "I29")
    set_widths(trend_ws, {"A": 22, "B": 12, "C": 12, "D": 12, "E": 12, "F": 12, "G": 12, "H": 4, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14, "O": 14, "P": 14, "Q": 14})

    setup_sheet(causal_ws, "Causal and Consequence Analysis", "Excel equivalents of cause, casualty, human-factor, and seasonal interaction views.")
    style_section_header(causal_ws, 4, "Causal Pareto View", width=6)
    cp_headers = ["Cause", "Incidents", "Casualties", "Fatalities", "Corridors Affected", "Cumulative %"]
    style_table_header(causal_ws, 5, 1, cp_headers)
    for idx, row in enumerate(insights["causeParetoData"], start=6):
        cause = row["cause"]
        causal_ws.cell(idx, 1, cause)
        causal_ws.cell(idx, 2, f'=COUNTIF(Enriched!$Q$5:$Q${raw_last},"{cause}")')
        causal_ws.cell(idx, 3, f'=SUMIF(Enriched!$Q$5:$Q${raw_last},"{cause}",Enriched!$W$5:$W${raw_last})')
        causal_ws.cell(idx, 4, f'=SUMIF(Enriched!$Q$5:$Q${raw_last},"{cause}",Enriched!$U$5:$U${raw_last})')
        causal_ws.cell(idx, 5, row["corridorsAffected"])
        causal_ws.cell(idx, 6, f'=ROUND(SUM($B$6:B{idx})/SUM($B$6:$B${5 + len(insights["causeParetoData"])})*100,0)')
    style_data_range(causal_ws, 5, 5 + len(insights["causeParetoData"]), 1, 6)
    pareto = BarChart()
    pareto.type = "col"
    pareto.title = "Causal Pareto View"
    pareto.height = 8
    pareto.width = 11.5
    pareto.add_data(Reference(causal_ws, min_col=2, max_col=2, min_row=5, max_row=5 + len(insights["causeParetoData"])), titles_from_data=True)
    pareto.set_categories(Reference(causal_ws, min_col=1, min_row=6, max_row=5 + len(insights["causeParetoData"])))
    create_chart_data_titles(pareto, "#2563EB")
    pareto_line = LineChart()
    pareto_line.y_axis.axId = 200
    pareto_line.y_axis.crosses = "max"
    pareto_line.y_axis.scaling.max = 100
    pareto_line.add_data(Reference(causal_ws, min_col=6, max_col=6, min_row=5, max_row=5 + len(insights["causeParetoData"])), titles_from_data=True)
    pareto_line.set_categories(Reference(causal_ws, min_col=1, min_row=6, max_row=5 + len(insights["causeParetoData"])))
    pareto_line.ser[0].graphicalProperties.line.solidFill = "991B1B"
    pareto_line.ser[0].marker.symbol = "circle"
    pareto += pareto_line
    causal_ws.add_chart(pareto, "H5")

    start = 22
    style_section_header(causal_ws, start, "Casualty Burden by Incident Type", width=6)
    cb_headers = ["Incident Type", "Incidents", "Fatalities", "Injuries", "Casualties"]
    style_table_header(causal_ws, start + 1, 1, cb_headers)
    for idx, row in enumerate(insights["casualtiesByType"], start=start + 2):
        incident_type = row["type"]
        causal_ws.cell(idx, 1, incident_type)
        causal_ws.cell(idx, 2, f'=COUNTIF(Enriched!$C$5:$C${raw_last},"{incident_type}")')
        causal_ws.cell(idx, 3, f'=SUMIF(Enriched!$C$5:$C${raw_last},"{incident_type}",Enriched!$U$5:$U${raw_last})')
        causal_ws.cell(idx, 4, f'=SUMIF(Enriched!$C$5:$C${raw_last},"{incident_type}",Enriched!$V$5:$V${raw_last})')
        causal_ws.cell(idx, 5, f"=C{idx}+D{idx}")
    style_data_range(causal_ws, start + 1, start + 1 + len(insights["casualtiesByType"]), 1, 5)
    cas_chart = BarChart()
    cas_chart.type = "bar"
    cas_chart.grouping = "stacked"
    cas_chart.overlap = 100
    cas_chart.title = "Casualty Burden by Incident Type"
    cas_chart.height = 8
    cas_chart.width = 11.5
    cas_chart.add_data(Reference(causal_ws, min_col=3, max_col=4, min_row=start + 1, max_row=start + 1 + len(insights["casualtiesByType"])), titles_from_data=True)
    cas_chart.set_categories(Reference(causal_ws, min_col=1, min_row=start + 2, max_row=start + 1 + len(insights["casualtiesByType"])))
    causal_ws.add_chart(cas_chart, "H23")

    start = 41
    style_section_header(causal_ws, start, "Human-Factor Signals", width=4)
    hf_headers = ["Behavior Factor", "Count", "Share %"]
    style_table_header(causal_ws, start + 1, 1, hf_headers)
    for idx, row in enumerate(insights["behaviorFactorBreakdown"], start=start + 2):
        factor = row["factor"]
        causal_ws.cell(idx, 1, factor)
        causal_ws.cell(idx, 2, f'=COUNTIF(Enriched!$R$5:$R${raw_last},"*{factor}*")')
        causal_ws.cell(idx, 3, f'=ROUND(B{idx}/COUNTA(Enriched!$A$5:$A${raw_last})*100,0)')
    style_data_range(causal_ws, start + 1, start + 1 + len(insights["behaviorFactorBreakdown"]), 1, 3)
    hf_chart = BarChart()
    hf_chart.type = "col"
    hf_chart.title = "Human-Factor Signals"
    hf_chart.height = 8
    hf_chart.width = 11.5
    hf_chart.add_data(Reference(causal_ws, min_col=2, max_col=2, min_row=start + 1, max_row=start + 1 + len(insights["behaviorFactorBreakdown"])), titles_from_data=True)
    hf_chart.set_categories(Reference(causal_ws, min_col=1, min_row=start + 2, max_row=start + 1 + len(insights["behaviorFactorBreakdown"])))
    causal_ws.add_chart(hf_chart, "H42")

    start = 60
    style_section_header(causal_ws, start, "Season and Visibility Interaction", width=6)
    causal_ws.cell(start + 1, 1, "Season")
    for offset, visibility in enumerate(VISIBILITY_ORDER, start=2):
        causal_ws.cell(start + 1, offset, visibility)
    style_table_header(causal_ws, start + 1, 1, ["Season", *VISIBILITY_ORDER])
    for row_offset, season in enumerate(SEASON_ORDER, start=start + 2):
        causal_ws.cell(row_offset, 1, season)
        for col_offset, visibility in enumerate(VISIBILITY_ORDER, start=2):
            causal_ws.cell(
                row_offset,
                col_offset,
                f'=SUMIFS(Enriched!$Y$5:$Y${raw_last},Enriched!$P$5:$P${raw_last},"{season}",Enriched!$L$5:$L${raw_last},"{visibility}")',
            )
    style_data_range(causal_ws, start + 1, start + 1 + len(SEASON_ORDER), 1, 1 + len(VISIBILITY_ORDER))
    matrix_range = f"B{start + 2}:E{start + 1 + len(SEASON_ORDER)}"
    causal_ws.conditional_formatting.add(
        matrix_range,
        ColorScaleRule(start_type="min", start_color="FEE2E2", mid_type="percentile", mid_value=50, mid_color="F87171", end_type="max", end_color="7F1D1D"),
    )
    causal_ws.cell(start + 8, 1, "Note")
    causal_ws.cell(start + 8, 1).fill = SUBSECTION_FILL
    causal_ws.cell(start + 8, 2, "The mockup used a matrix here because seasonal pressure by visibility band is easier to scan than another line or bar chart.")
    causal_ws.cell(start + 8, 2).alignment = Alignment(wrap_text=True)
    set_widths(causal_ws, {"A": 22, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 4, "H": 14, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14, "O": 14, "P": 14, "Q": 14})

    setup_sheet(exposure_ws, "Maritime Exposure Analysis", "Corridor, sea area, vessel class, and seasonal operating pressure views.")
    start = 4
    style_section_header(exposure_ws, start, "Corridor Risk Ranking", width=7)
    corr_headers = ["Corridor", "Sea Area", "Records", "Fatalities", "Injuries", "Risk Index", "Avg. Completeness"]
    style_table_header(exposure_ws, start + 1, 1, corr_headers)
    for idx, row in enumerate(insights["corridorHotspots"], start=start + 2):
        corridor = row["corridor"]
        exposure_ws.cell(idx, 1, corridor)
        exposure_ws.cell(idx, 2, row["seaArea"])
        exposure_ws.cell(idx, 3, f'=COUNTIF(Enriched!$J$5:$J${raw_last},"{corridor}")')
        exposure_ws.cell(idx, 4, f'=SUMIF(Enriched!$J$5:$J${raw_last},"{corridor}",Enriched!$U$5:$U${raw_last})')
        exposure_ws.cell(idx, 5, f'=SUMIF(Enriched!$J$5:$J${raw_last},"{corridor}",Enriched!$V$5:$V${raw_last})')
        exposure_ws.cell(idx, 6, f'=SUMIF(Enriched!$J$5:$J${raw_last},"{corridor}",Enriched!$Y$5:$Y${raw_last})')
        exposure_ws.cell(idx, 7, f'=ROUND(AVERAGEIF(Enriched!$J$5:$J${raw_last},"{corridor}",Enriched!$AD$5:$AD${raw_last}),0)')
    style_data_range(exposure_ws, start + 1, start + 1 + len(insights["corridorHotspots"]), 1, 7)
    corr_chart = BarChart()
    corr_chart.type = "bar"
    corr_chart.title = "Corridor Risk Ranking"
    corr_chart.height = 8
    corr_chart.width = 11.5
    corr_chart.add_data(Reference(exposure_ws, min_col=6, max_col=6, min_row=start + 1, max_row=start + 1 + min(6, len(insights["corridorHotspots"]))), titles_from_data=True)
    corr_chart.set_categories(Reference(exposure_ws, min_col=1, min_row=start + 2, max_row=start + 1 + min(6, len(insights["corridorHotspots"]))))
    exposure_ws.add_chart(corr_chart, "I5")

    start = 24
    style_section_header(exposure_ws, start, "Sea Area Exposure", width=5)
    sea_headers = ["Sea Area", "Records", "Fatalities", "Injuries", "Risk Index"]
    style_table_header(exposure_ws, start + 1, 1, sea_headers)
    for idx, row in enumerate(insights["seaAreaBreakdown"], start=start + 2):
        sea_area = row["seaArea"]
        exposure_ws.cell(idx, 1, sea_area)
        exposure_ws.cell(idx, 2, f'=COUNTIF(Enriched!$I$5:$I${raw_last},"{sea_area}")')
        exposure_ws.cell(idx, 3, f'=SUMIF(Enriched!$I$5:$I${raw_last},"{sea_area}",Enriched!$U$5:$U${raw_last})')
        exposure_ws.cell(idx, 4, f'=SUMIF(Enriched!$I$5:$I${raw_last},"{sea_area}",Enriched!$V$5:$V${raw_last})')
        exposure_ws.cell(idx, 5, f'=SUMIF(Enriched!$I$5:$I${raw_last},"{sea_area}",Enriched!$Y$5:$Y${raw_last})')
    style_data_range(exposure_ws, start + 1, start + 1 + len(insights["seaAreaBreakdown"]), 1, 5)
    sea_chart = BarChart()
    sea_chart.type = "bar"
    sea_chart.title = "Sea Area Exposure"
    sea_chart.height = 8
    sea_chart.width = 11.5
    sea_chart.add_data(Reference(exposure_ws, min_col=5, max_col=5, min_row=start + 1, max_row=start + 1 + len(insights["seaAreaBreakdown"])), titles_from_data=True)
    sea_chart.set_categories(Reference(exposure_ws, min_col=1, min_row=start + 2, max_row=start + 1 + len(insights["seaAreaBreakdown"])))
    exposure_ws.add_chart(sea_chart, "I25")

    start = 43
    style_section_header(exposure_ws, start, "Vessel Class Exposure", width=6)
    ve_headers = ["Vessel Type", "Record Count", "Risk Index", "Casualties", "Severity Index"]
    style_table_header(exposure_ws, start + 1, 1, ve_headers)
    for idx, row in enumerate(insights["vesselExposureData"], start=start + 2):
        vessel = row["type"]
        exposure_ws.cell(idx, 1, vessel)
        exposure_ws.cell(idx, 2, f'=COUNTIF(Enriched!$E$5:$E${raw_last},"{vessel}")')
        exposure_ws.cell(idx, 3, f'=SUMIF(Enriched!$E$5:$E${raw_last},"{vessel}",Enriched!$Y$5:$Y${raw_last})')
        exposure_ws.cell(idx, 4, f'=SUMIF(Enriched!$E$5:$E${raw_last},"{vessel}",Enriched!$W$5:$W${raw_last})')
        exposure_ws.cell(idx, 5, f'=ROUND(C{idx}/MAX(B{idx},1),0)')
    style_data_range(exposure_ws, start + 1, start + 1 + len(insights["vesselExposureData"]), 1, 5)
    vessel_chart = BarChart()
    vessel_chart.type = "bar"
    vessel_chart.title = "Vessel Class Exposure"
    vessel_chart.height = 8
    vessel_chart.width = 11.5
    vessel_chart.add_data(Reference(exposure_ws, min_col=3, max_col=3, min_row=start + 1, max_row=start + 1 + len(insights["vesselExposureData"])), titles_from_data=True)
    vessel_chart.set_categories(Reference(exposure_ws, min_col=1, min_row=start + 2, max_row=start + 1 + len(insights["vesselExposureData"])))
    exposure_ws.add_chart(vessel_chart, "I44")

    start = 62
    style_section_header(exposure_ws, start, "Seasonal Operating Pressure", width=5)
    so_headers = ["Season", "Incidents", "Casualties", "Risk Index"]
    style_table_header(exposure_ws, start + 1, 1, so_headers)
    for idx, row in enumerate(insights["seasonBreakdown"], start=start + 2):
        season = row["season"]
        exposure_ws.cell(idx, 1, season)
        exposure_ws.cell(idx, 2, f'=COUNTIF(Enriched!$P$5:$P${raw_last},"{season}")')
        exposure_ws.cell(idx, 3, f'=SUMIF(Enriched!$P$5:$P${raw_last},"{season}",Enriched!$W$5:$W${raw_last})')
        exposure_ws.cell(idx, 4, f'=SUMIF(Enriched!$P$5:$P${raw_last},"{season}",Enriched!$Y$5:$Y${raw_last})')
    style_data_range(exposure_ws, start + 1, start + 1 + len(insights["seasonBreakdown"]), 1, 4)
    season_chart = BarChart()
    season_chart.type = "col"
    season_chart.title = "Seasonal Operating Pressure"
    season_chart.height = 8
    season_chart.width = 11.5
    season_chart.add_data(Reference(exposure_ws, min_col=2, max_col=2, min_row=start + 1, max_row=start + 1 + len(insights["seasonBreakdown"])), titles_from_data=True)
    season_chart.set_categories(Reference(exposure_ws, min_col=1, min_row=start + 2, max_row=start + 1 + len(insights["seasonBreakdown"])))
    season_line = LineChart()
    season_line.y_axis.axId = 201
    season_line.y_axis.crosses = "max"
    season_line.add_data(Reference(exposure_ws, min_col=3, max_col=3, min_row=start + 1, max_row=start + 1 + len(insights["seasonBreakdown"])), titles_from_data=True)
    season_line.set_categories(Reference(exposure_ws, min_col=1, min_row=start + 2, max_row=start + 1 + len(insights["seasonBreakdown"])))
    season_line.ser[0].graphicalProperties.line.solidFill = "991B1B"
    season_line.ser[0].marker.symbol = "circle"
    season_chart += season_line
    exposure_ws.add_chart(season_chart, "I63")

    start = 81
    style_section_header(exposure_ws, start, "Priority Corridor Review Table", width=7)
    style_table_header(exposure_ws, start + 1, 1, ["Sea Area", "Corridor", "Records", "Fatalities", "Risk Index", "Dominant Cause", "Avg. Completeness"])
    for idx, row in enumerate(insights["corridorInvestigationTable"], start=start + 2):
        exposure_ws.cell(idx, 1, row["seaArea"])
        exposure_ws.cell(idx, 2, row["corridor"])
        exposure_ws.cell(idx, 3, f'=COUNTIF(Enriched!$J$5:$J${raw_last},"{row["corridor"]}")')
        exposure_ws.cell(idx, 4, f'=SUMIF(Enriched!$J$5:$J${raw_last},"{row["corridor"]}",Enriched!$U$5:$U${raw_last})')
        exposure_ws.cell(idx, 5, f'=SUMIF(Enriched!$J$5:$J${raw_last},"{row["corridor"]}",Enriched!$Y$5:$Y${raw_last})')
        exposure_ws.cell(idx, 6, row["dominantCause"])
        exposure_ws.cell(idx, 7, f'=ROUND(AVERAGEIF(Enriched!$J$5:$J${raw_last},"{row["corridor"]}",Enriched!$AD$5:$AD${raw_last}),0)')
    style_data_range(exposure_ws, start + 1, start + 1 + len(insights["corridorInvestigationTable"]), 1, 7)
    set_widths(exposure_ws, {"A": 22, "B": 28, "C": 12, "D": 12, "E": 12, "F": 22, "G": 18, "H": 4, "I": 14, "J": 14, "K": 14, "L": 14, "M": 14, "N": 14, "O": 14, "P": 14, "Q": 14})

    setup_sheet(quality_ws, "Evidence Quality Analysis", "Completeness, workflow maturity, and reporting-channel views.")
    start = 4
    style_section_header(quality_ws, start, "SAFERSEAS Intake Completeness", width=5)
    dc_headers = ["Metric", "Description", "Captured", "Total", "Coverage %"]
    style_table_header(quality_ws, start + 1, 1, dc_headers)
    completeness_formulas = [
        ('Registry identifiers', "IMO number or official number captured", '=COUNTIF(Enriched!$Z$5:$Z$999,1)'),
        ('Voyage and route details', "Origin, destination, and route context documented", '=COUNTIFS(Enriched!$AA$5:$AA$999,1,Enriched!$AC$5:$AC$999,"<>Route pending")'),
        ('Weather observations', "Wave height, wind force, and visibility present", '=COUNTIF(Enriched!$AB$5:$AB$999,1)'),
        ('Cause findings', "Primary cause and findings documented", '=COUNTIFS(Raw_Incidents!$AR$5:$AR$999,"<>",Raw_Incidents!$AS$5:$AS$999,"<>")'),
        ('Linked evidence', "Attachments or supporting documents referenced", '=COUNTIF(Enriched!$AE$5:$AE$999,1)'),
        ('Ownership and operator', "Accountable entity information available", '=COUNTIFS(Raw_Incidents!$AB$5:$AB$999,"<>")+COUNTIFS(Raw_Incidents!$AD$5:$AD$999,"<>")-COUNTIFS(Raw_Incidents!$AB$5:$AB$999,"<>",Raw_Incidents!$AD$5:$AD$999,"<>")'),
    ]
    for idx, (label, desc, captured_formula) in enumerate(completeness_formulas, start=start + 2):
        quality_ws.cell(idx, 1, label)
        quality_ws.cell(idx, 2, desc)
        quality_ws.cell(idx, 3, captured_formula)
        quality_ws.cell(idx, 4, "=COUNTA(Enriched!$A$5:$A$999)")
        quality_ws.cell(idx, 5, f'=ROUND(C{idx}/D{idx}*100,0)')
    style_data_range(quality_ws, start + 1, start + 1 + len(completeness_formulas), 1, 5)
    quality_ws.conditional_formatting.add(
        f"E{start + 2}:E{start + 1 + len(completeness_formulas)}",
        DataBarRule(start_type="num", start_value=0, end_type="num", end_value=100, color="2563EB", showValue=True),
    )

    start = 4
    style_section_header(quality_ws, start, "Critical Gaps for Policy Use", width=5)
    gap_col = 8
    style_table_header(quality_ws, start + 1, gap_col, ["Metric", "Description", "Coverage %"])
    for idx, row in enumerate(insights["criticalGapData"], start=start + 2):
        label = row["label"]
        quality_ws.cell(idx, gap_col, label)
        quality_ws.cell(idx, gap_col + 1, row["description"])
        lookup_range = f"$A${start + 2}:$E${start + 1 + len(completeness_formulas)}"
        quality_ws.cell(idx, gap_col + 2, f'=INDEX($E${start + 2}:$E${start + 1 + len(completeness_formulas)},MATCH(H{idx},$A${start + 2}:$A${start + 1 + len(completeness_formulas)},0))')
    style_data_range(quality_ws, start + 1, start + 1 + len(insights["criticalGapData"]), gap_col, gap_col + 2)
    gap_chart = BarChart()
    gap_chart.type = "bar"
    gap_chart.title = "Critical Gaps for Policy Use"
    gap_chart.height = 7.5
    gap_chart.width = 10
    gap_chart.add_data(Reference(quality_ws, min_col=10, max_col=10, min_row=5, max_row=5 + len(insights["criticalGapData"])), titles_from_data=True)
    gap_chart.set_categories(Reference(quality_ws, min_col=8, min_row=6, max_row=5 + len(insights["criticalGapData"])))
    quality_ws.add_chart(gap_chart, "L5")

    start = 23
    style_section_header(quality_ws, start, "Workflow and Record Quality", width=4)
    style_table_header(quality_ws, start + 1, 1, ["Status", "Records", "Avg. Completeness"])
    for idx, row in enumerate(insights["statusCompletenessData"], start=start + 2):
        status = row["status"]
        quality_ws.cell(idx, 1, status)
        quality_ws.cell(idx, 2, f'=COUNTIF(Enriched!$G$5:$G${raw_last},"{status}")')
        quality_ws.cell(idx, 3, f'=ROUND(AVERAGEIF(Enriched!$G$5:$G${raw_last},"{status}",Enriched!$AD$5:$AD${raw_last}),0)')
    style_data_range(quality_ws, start + 1, start + 1 + len(insights["statusCompletenessData"]), 1, 3)
    status_chart = BarChart()
    status_chart.type = "col"
    status_chart.title = "Workflow and Record Quality"
    status_chart.height = 7.5
    status_chart.width = 10
    status_chart.add_data(Reference(quality_ws, min_col=2, max_col=2, min_row=start + 1, max_row=start + 1 + len(insights["statusCompletenessData"])), titles_from_data=True)
    status_chart.set_categories(Reference(quality_ws, min_col=1, min_row=start + 2, max_row=start + 1 + len(insights["statusCompletenessData"])))
    status_line = LineChart()
    status_line.y_axis.axId = 201
    status_line.y_axis.crosses = "max"
    status_line.add_data(Reference(quality_ws, min_col=3, max_col=3, min_row=start + 1, max_row=start + 1 + len(insights["statusCompletenessData"])), titles_from_data=True)
    status_line.set_categories(Reference(quality_ws, min_col=1, min_row=start + 2, max_row=start + 1 + len(insights["statusCompletenessData"])))
    status_line.ser[0].graphicalProperties.line.solidFill = "0F766E"
    status_line.ser[0].marker.symbol = "diamond"
    status_chart += status_line
    quality_ws.add_chart(status_chart, "E24")

    start = 23
    style_section_header(quality_ws, start, "Reporting Pathways", width=4)
    rp_col = 12
    style_table_header(quality_ws, start + 1, rp_col, ["Reporting Method", "Records"])
    for idx, row in enumerate(insights["reportingMethodBreakdown"], start=start + 2):
        method = row["method"]
        quality_ws.cell(idx, rp_col, method)
        quality_ws.cell(idx, rp_col + 1, f'=COUNTIF(Enriched!$AK$5:$AK${raw_last},"{method}")')
    style_data_range(quality_ws, start + 1, start + 1 + len(insights["reportingMethodBreakdown"]), rp_col, rp_col + 1)
    set_widths(quality_ws, {"A": 24, "B": 14, "C": 18, "D": 4, "E": 14, "F": 14, "G": 14, "H": 22, "I": 30, "J": 12, "K": 4, "L": 22, "M": 12, "N": 14, "O": 14, "P": 14, "Q": 14})

    setup_sheet(policy_ws, "Policy Brief and Evidence Packets", "Excel equivalent of the mockup's policy page, grounded in formula-driven evidence tables.")
    add_card(policy_ws, "A4", "Top Corridor at Risk", "=Exposure_Analysis!A6", "Highest-risk corridor from the exposure sheet", "991B1B")
    add_card(policy_ws, "C4", "Leading Repeat Cause", "=Causal_Analysis!A6", "Top cause from the Pareto table", "1D4ED8")
    add_card(policy_ws, "E4", "Small-Craft Weather Records", '=COUNTIFS(Enriched!$Q$5:$Q$999,"Weather Conditions",Enriched!$E$5:$E$999,"Fishing Vessel")', "Fishing-vessel weather cases in the model", "D97706")
    add_card(policy_ws, "G4", "Avg. Evidence Completeness", '=ROUND(AVERAGE(Enriched!$AD$5:$AD$999),0)&"%"', "Overall completeness across enriched records", "7C3AED")

    style_section_header(policy_ws, 11, "Evidence Packages for Immediate Action", width=12)
    packet_headers = [
        "Title", "Priority", "Instrument", "Lead", "Supporting Records", "Fatalities", "Injuries",
        "Confidence Score", "Confidence Label", "Evidence Summary", "Confidence Note", "Recommended Action",
    ]
    style_table_header(policy_ws, 12, 1, packet_headers)
    packet_rows = insights["policyEvidencePackets"]
    for idx, packet in enumerate(packet_rows, start=13):
        policy_ws.cell(idx, 1, packet["title"])
        policy_ws.cell(idx, 2, packet["priority"])
        policy_ws.cell(idx, 3, packet["policyInstrument"])
        policy_ws.cell(idx, 4, packet["lead"])

        if idx == 13:
            corridor = packet["corridors"][0]
            policy_ws.cell(idx, 5, f'=COUNTIF(Enriched!$J$5:$J${raw_last},"{corridor}")')
            policy_ws.cell(idx, 6, f'=SUMIF(Enriched!$J$5:$J${raw_last},"{corridor}",Enriched!$U$5:$U${raw_last})')
            policy_ws.cell(idx, 7, f'=SUMIF(Enriched!$J$5:$J${raw_last},"{corridor}",Enriched!$V$5:$V${raw_last})')
            policy_ws.cell(idx, 8, f'=ROUND((AVERAGEIF(Enriched!$J$5:$J${raw_last},"{corridor}",Enriched!$AD$5:$AD${raw_last})*0.5)+(COUNTIFS(Enriched!$J$5:$J${raw_last},"{corridor}",Enriched!$AE$5:$AE${raw_last},1)/E{idx}*100*0.25)+(COUNTIFS(Enriched!$J$5:$J${raw_last},"{corridor}",Enriched!$AF$5:$AF${raw_last},1)/E{idx}*100*0.25),0)')
        elif idx == 14:
            policy_ws.cell(idx, 5, f'=COUNTIF(Enriched!$Q$5:$Q${raw_last},"Navigation Error")')
            policy_ws.cell(idx, 6, f'=SUMIF(Enriched!$Q$5:$Q${raw_last},"Navigation Error",Enriched!$U$5:$U${raw_last})')
            policy_ws.cell(idx, 7, f'=SUMIF(Enriched!$Q$5:$Q${raw_last},"Navigation Error",Enriched!$V$5:$V${raw_last})')
            policy_ws.cell(idx, 8, f'=ROUND((AVERAGEIF(Enriched!$Q$5:$Q${raw_last},"Navigation Error",Enriched!$AD$5:$AD${raw_last})*0.5)+(COUNTIFS(Enriched!$Q$5:$Q${raw_last},"Navigation Error",Enriched!$AE$5:$AE${raw_last},1)/E{idx}*100*0.25)+(COUNTIFS(Enriched!$Q$5:$Q${raw_last},"Navigation Error",Enriched!$AF$5:$AF${raw_last},1)/E{idx}*100*0.25),0)')
        elif idx == 15:
            policy_ws.cell(idx, 5, f'=COUNTIFS(Enriched!$Q$5:$Q${raw_last},"Weather Conditions",Enriched!$E$5:$E${raw_last},"Fishing Vessel")')
            policy_ws.cell(idx, 6, f'=SUMIFS(Enriched!$U$5:$U${raw_last},Enriched!$Q$5:$Q${raw_last},"Weather Conditions",Enriched!$E$5:$E${raw_last},"Fishing Vessel")')
            policy_ws.cell(idx, 7, f'=SUMIFS(Enriched!$V$5:$V${raw_last},Enriched!$Q$5:$Q${raw_last},"Weather Conditions",Enriched!$E$5:$E${raw_last},"Fishing Vessel")')
            policy_ws.cell(idx, 8, f'=ROUND((AVERAGEIFS(Enriched!$AD$5:$AD${raw_last},Enriched!$Q$5:$Q${raw_last},"Weather Conditions",Enriched!$E$5:$E${raw_last},"Fishing Vessel")*0.5)+(COUNTIFS(Enriched!$Q$5:$Q${raw_last},"Weather Conditions",Enriched!$E$5:$E${raw_last},"Fishing Vessel",Enriched!$AE$5:$AE${raw_last},1)/E{idx}*100*0.25)+(COUNTIFS(Enriched!$Q$5:$Q${raw_last},"Weather Conditions",Enriched!$E$5:$E${raw_last},"Fishing Vessel",Enriched!$AF$5:$AF${raw_last},1)/E{idx}*100*0.25),0)')
        else:
            policy_ws.cell(idx, 5, f'=COUNTIF(Enriched!$Q$5:$Q${raw_last},"Mechanical Failure")')
            policy_ws.cell(idx, 6, f'=SUMIF(Enriched!$Q$5:$Q${raw_last},"Mechanical Failure",Enriched!$U$5:$U${raw_last})')
            policy_ws.cell(idx, 7, f'=SUMIF(Enriched!$Q$5:$Q${raw_last},"Mechanical Failure",Enriched!$V$5:$V${raw_last})')
            policy_ws.cell(idx, 8, f'=ROUND((AVERAGEIF(Enriched!$Q$5:$Q${raw_last},"Mechanical Failure",Enriched!$AD$5:$AD${raw_last})*0.5)+(COUNTIFS(Enriched!$Q$5:$Q${raw_last},"Mechanical Failure",Enriched!$AE$5:$AE${raw_last},1)/E{idx}*100*0.25)+(COUNTIFS(Enriched!$Q$5:$Q${raw_last},"Mechanical Failure",Enriched!$AF$5:$AF${raw_last},1)/E{idx}*100*0.25),0)')

        policy_ws.cell(idx, 9, f'=IF(H{idx}>=80,"High",IF(H{idx}>=65,"Moderate","Low"))')
        policy_ws.cell(idx, 10, f'=E{idx}&" supporting records, "&F{idx}&" fatalities, "&G{idx}&" injuries, and "&ROUND(AVERAGE(Enriched!$AD$5:$AD${raw_last}),0)&"% average record completeness."')
        policy_ws.cell(idx, 11, f'=I{idx}&" confidence based on subset completeness, linked evidence coverage, and review maturity."')
        policy_ws.cell(idx, 12, packet["recommendedAction"])
    style_data_range(policy_ws, 12, 12 + len(packet_rows), 1, 12)

    corr_chart_copy = BarChart()
    corr_chart_copy.type = "bar"
    corr_chart_copy.title = "Traceable Corridor Evidence"
    corr_chart_copy.height = 7.5
    corr_chart_copy.width = 10
    corr_chart_copy.add_data(Reference(exposure_ws, min_col=6, max_col=6, min_row=5, max_row=11), titles_from_data=True)
    corr_chart_copy.set_categories(Reference(exposure_ws, min_col=1, min_row=6, max_row=11))
    policy_ws.add_chart(corr_chart_copy, "N12")

    cause_chart_copy = BarChart()
    cause_chart_copy.type = "col"
    cause_chart_copy.title = "Repeat Cause Burden"
    cause_chart_copy.height = 7.5
    cause_chart_copy.width = 10
    cause_chart_copy.add_data(Reference(causal_ws, min_col=2, max_col=2, min_row=5, max_row=10), titles_from_data=True)
    cause_chart_copy.set_categories(Reference(causal_ws, min_col=1, min_row=6, max_row=10))
    policy_ws.add_chart(cause_chart_copy, "N29")

    start = 29
    style_section_header(policy_ws, start, "Corridor Evidence Register", width=7)
    style_table_header(policy_ws, start + 1, 1, ["Sea Area", "Corridor", "Records", "Fatalities", "Risk Index", "Dominant Cause", "Avg. Completeness"])
    for idx, row in enumerate(insights["corridorInvestigationTable"], start=start + 2):
        policy_ws.cell(idx, 1, row["seaArea"])
        policy_ws.cell(idx, 2, row["corridor"])
        policy_ws.cell(idx, 3, f'=COUNTIF(Enriched!$J$5:$J${raw_last},"{row["corridor"]}")')
        policy_ws.cell(idx, 4, f'=SUMIF(Enriched!$J$5:$J${raw_last},"{row["corridor"]}",Enriched!$U$5:$U${raw_last})')
        policy_ws.cell(idx, 5, f'=SUMIF(Enriched!$J$5:$J${raw_last},"{row["corridor"]}",Enriched!$Y$5:$Y${raw_last})')
        policy_ws.cell(idx, 6, row["dominantCause"])
        policy_ws.cell(idx, 7, f'=ROUND(AVERAGEIF(Enriched!$J$5:$J${raw_last},"{row["corridor"]}",Enriched!$AD$5:$AD${raw_last}),0)')
    style_data_range(policy_ws, start + 1, start + 1 + len(insights["corridorInvestigationTable"]), 1, 7)

    style_section_header(policy_ws, 45, "Confidence and Data Caveats", width=6)
    style_table_header(policy_ws, 46, 1, ["Metric", "Coverage %", "Description"])
    for idx, row in enumerate(insights["criticalGapData"][:3], start=47):
        policy_ws.cell(idx, 1, row["label"])
        policy_ws.cell(idx, 2, f'=INDEX(Quality_Analysis!$J$6:$J$11,MATCH(A{idx},Quality_Analysis!$H$6:$H$11,0))')
        policy_ws.cell(idx, 3, row["description"])
    style_data_range(policy_ws, 46, 49, 1, 3)
    policy_ws["A51"] = "Methodology Notes"
    policy_ws["A51"].fill = SECTION_FILL
    policy_ws["A51"].font = Font(color="1E3A8A", bold=True)
    policy_ws["A52"] = (
        "Risk Index = severity weight × max(casualty burden, 1). It is a prioritization proxy used by the mockup, not a validated official doctrine metric."
    )
    policy_ws["A53"] = (
        "Confidence Score blends subset completeness, linked evidence presence, and review maturity. It helps a developer implement the mockup, but should be validated by policy and investigation stakeholders."
    )
    policy_ws["A52"].alignment = Alignment(wrap_text=True)
    policy_ws["A53"].alignment = Alignment(wrap_text=True)
    set_widths(policy_ws, {"A": 28, "B": 16, "C": 18, "D": 24, "E": 16, "F": 12, "G": 12, "H": 14, "I": 16, "J": 34, "K": 32, "L": 40, "M": 4, "N": 14, "O": 14, "P": 14, "Q": 14, "R": 14, "S": 14, "T": 14, "U": 14})

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False

    return wb


def recalculate_with_excel(path: Path):
    import pythoncom  # type: ignore
    import pywintypes  # type: ignore
    import win32com.client  # type: ignore

    excel = None
    workbook = None

    pythoncom.CoInitialize()
    try:
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        excel.ScreenUpdating = False
        excel.EnableEvents = False
        excel.AskToUpdateLinks = False

        open_error = None
        for attempt in range(6):
            try:
                workbook = excel.Workbooks.Open(
                    str(path.resolve()),
                    UpdateLinks=0,
                    ReadOnly=False,
                    IgnoreReadOnlyRecommended=True,
                    AddToMru=False,
                )
                time.sleep(1)
                open_error = None
                break
            except pywintypes.com_error as exc:
                open_error = exc
                time.sleep(attempt + 1)

        if workbook is None:
            raise open_error or RuntimeError(f"Unable to open workbook for recalculation: {path}")

        calc_error = None
        for attempt in range(6):
            try:
                workbook.Application.CalculateFullRebuild()
                time.sleep(1)
                workbook.Save()
                calc_error = None
                break
            except pywintypes.com_error as exc:
                calc_error = exc
                time.sleep(attempt + 1)

        if calc_error is not None:
            raise calc_error
    finally:
        if workbook is not None:
            try:
                workbook.Close(SaveChanges=True)
            except Exception:
                pass
        if excel is not None:
            try:
                excel.Quit()
            except Exception:
                pass
        pythoncom.CoUninitialize()


def verify_workbook(path: Path, insights: dict):
    wb = load_workbook(path, data_only=True)
    trend = wb["Trend_Summary"]
    exposure = wb["Exposure_Analysis"]
    quality = wb["Quality_Analysis"]
    policy = wb["Policy_Brief"]

    assert trend["A4"].value == "Incident Records"
    assert trend["A5"].value == insights["analyticsSummary"]["totalIncidents"]
    assert trend["C5"].value == insights["analyticsSummary"]["highConsequenceCount"]
    assert trend["E5"].value == len(insights["corridorHotspots"])
    assert str(trend["G5"].value).replace("%", "") == str(insights["analyticsSummary"]["linkedEvidenceCoverage"])

    first_month = insights["monthlyTrendData"][0]
    assert trend["A13"].value == first_month["period"]
    assert trend["B13"].value == first_month["incidents"]
    assert trend["C13"].value == first_month["fatalities"]
    assert trend["D13"].value == first_month["injuries"]

    first_corridor = insights["corridorHotspots"][0]
    assert exposure["A6"].value == first_corridor["corridor"]
    assert exposure["F6"].value == first_corridor["riskIndex"]

    first_gap = insights["criticalGapData"][0]
    assert quality["H6"].value == first_gap["label"]
    assert quality["J6"].value == first_gap["percentage"]

    first_packet = insights["policyEvidencePackets"][0]
    assert policy["A13"].value == first_packet["title"]
    assert policy["E13"].value == first_packet["supportingRecords"]
    assert policy["F13"].value == first_packet["fatalities"]
    assert policy["G13"].value == first_packet["injuries"]


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with tempfile.TemporaryDirectory() as tmp:
        temp_dir = Path(tmp)
        mock_incidents, insights = run_export(temp_dir)
        workbook = build_workbook(mock_incidents, insights)
        workbook.save(OUTPUT_PATH)
        recalculate_with_excel(OUTPUT_PATH)
        verify_workbook(OUTPUT_PATH, insights)

    print(OUTPUT_PATH)


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:  # pragma: no cover
        print(f"Workbook build failed: {exc}", file=sys.stderr)
        raise
